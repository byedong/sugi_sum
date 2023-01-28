import os
import shutil
import pandas as pd
from openpyxl import load_workbook, Workbook
from distutils.dir_util import copy_tree
from operator import index
import tkinter
from tkinter import filedialog

#FutureWaring 없에기
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.filterwarnings(action='ignore')
#####################################################################

root = tkinter.Tk()
root.withdraw()
basepath = filedialog.askdirectory(parent=root,initialdir="/",title='Please select a directory')

# scandir()를 사용하여 모든 서브디렉토리 리스트하기

#basepath = './'

#print(basepath)

print('######################')
print('#   수기조사 통합중    #')
print('#   창을 닫지 마세요.! #')
print('######################')

# 하위폴더를 저장할 리스트 선언
dir_list = [] # 폴더명 + 파일명 ex) '2022년 05월31일 14시38분02초/수기조사서.xlsx'
folder_list = [] # 폴더명만 저장  ex) '2022년 05월31일 14시38분02초'
pic_list = [] # 사진폴더명 저장, ex) ' 2022년 05월31일 14시38분02초/사진 '
CONTENTS =[]
outfile_name = "수기조사서.xlsx"

# 폴더명 + 파일을 리스트에 저장 (폴더명\수기조사서.xlsx)
with os.scandir(basepath) as entries:
    for entry in entries:
        if entry.is_dir():
            folder_list.append(entry.name) # 폴더 이름저장
            temp = entry.name
            temp += '/수기조사서.xlsx'

            pic_temp = entry.name
            pic_temp += '/사진'
            dir_list.append(basepath + '/' +temp) # 폴더+이름 리스트로 저장
            pic_list.append(basepath + '/' +pic_temp) #폴더 + 사진 리스트로 저장
            #print(entry.name)


merge_df = pd.DataFrame()

# 파일을 열어서 추가

for file_name in dir_list:
    file_df = pd.read_excel(file_name, skiprows=4)
    columns = list(file_df.columns)
    temp_df = pd.DataFrame(file_df, columns=columns)
    # 기존꺼에서 추가 하기 apeend가 없어질 예정, 새로운 버전에서.
    merge_df = merge_df.append(temp_df, ignore_index=False)
    merge_df.to_excel("수기조사서.xlsx", index=False)


#엑셀파일저장
wb = load_workbook("수기조사서.xlsx")
ws = wb.active
ws.insert_rows(1,4) #8번째에 row 한줄 추가
wb.save('수기조사서.xlsx')

#print(folder_list)
#print(dir_list)
#print(pic_list)

#바탕화면에 "수기조사"라는 폴더 만들어서, 압축파일 및 "통합 폴더" 만들기
windows_user_name = os.path.expanduser('~')

#폴더생성
os.makedirs(f'{windows_user_name}//Desktop//수기_통합')
os.makedirs(f'{windows_user_name}//Desktop//수기_통합//통합')
os.makedirs(f'{windows_user_name}//Desktop//수기_통합//통합//사진')

sugi_dir = os.path.expanduser('~\\Desktop\\수기_통합')
sum_dir = os.path.expanduser('~\\Desktop\\수기_통합\\통합')
sum_pic_dir = os.path.expanduser('~\\Desktop\\수기_통합\\통합\\사진')

#수기조사 최종 폴더로 이동
shutil.move('수기조사서.xlsx', sugi_dir)

#######################################################################

df = pd.read_excel(sugi_dir + '/수기조사서.xlsx', skiprows=4)##경로 한번 확인 해얄듯
#print(df)
df.set_index("No")
df.apply(lambda x: x.str.strip(), axis = 1) ##공백제거가 안됨요!!!!!!!!!!

#품명에 "컴퓨터", "모니터"만 빼서 export_df에 저장
filter_list = ['컴퓨터', '모니터', '컴퓨터(업무)','컴퓨터(학급)']
export_df =df[df['품명'].isin(filter_list)] #컴퓨터, 모니터만 뺴기
export_df.set_index("No")

sorted_df = export_df.sort_values(by=["제조일자","모델명", "제조사"], axis = 0, ascending = [False, True, True])
sorted_df2 = sorted_df.sort_values(by=["품명","제조사"], axis = 0)

sorted_df2.to_excel(sugi_dir +'/모니터-컴퓨터.xlsx', index = None)

#수기조사서에서 "컴퓨터", "모니터" 제거 후, 저장
df = df[df.품명 != '컴퓨터']
df = df[df.품명 != '모니터']
sorted_df2 = df.sort_values("품명", ascending = [True])
sorted_df2.to_excel(sum_dir + '/수기조사서.xlsx', index = None)

##########################################################################
#shutil.move('수기조사서.xlsx', sum_dir)

#사진복사
for i in pic_list:
   temp = str(i)
   #print(temp)
   copy_tree(temp, sum_pic_dir)

#수기조사 하기전 위에 빈칸 4줄 넣기
wb = load_workbook(sum_dir + '/수기조사서.xlsx')
ws = wb.active
ws.insert_rows(1,4) #8번째에 row 한줄 추가
ws.merge_cells("A2:L3")
wb.save(sum_dir + '/수기조사서.xlsx')


#zip파일로 압축
os.chdir(sum_dir)
target_dir = os.getcwd()
result_dir = sugi_dir
os.chdir(result_dir)
#print(result_dir)
shutil.make_archive('통합','zip',target_dir)

os.remove(sugi_dir + '/수기조사서.xlsx')